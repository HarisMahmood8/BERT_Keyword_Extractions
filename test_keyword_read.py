import openpyxl


# Get dictionary of keywords and topics
def read_keyword_topics(sheet):
    res_dict = {}
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
        row_items = []
        for cell in row:
            row_items.append(cell)

        key, value = row_items[2], row_items[0]
        if key not in res_dict.keys():
            res_dict[key] = []
        res_dict[key].append(value)

    return res_dict


def export_to_excel(data, file_name):
    res_workbook = openpyxl.Workbook()
    sheet = res_workbook.active
    sheet.title = file_name

    sheet.append(["Category", "Sentiment", "Paragraph", "Keywords"])
    for row_data in data:
        sheet.append(row_data)

    res_workbook.save(file_name)


if __name__ == "__main__":
    # Functional test
    workbook = openpyxl.load_workbook("keywords_topics.xlsx")

    res_dict = read_keyword_topics(workbook["Key Words_Topics"])
    for key in res_dict:
        for item in res_dict[key]:
            print(key, item)

    data = [
        ["Name", "Age", "City", 1],
        ["Alice", 25, "New York", 2],
        ["Bob", 30, "Los Angeles", 3],
        ["Charlie", 22, "Chicago", 4],
    ]

    export_to_excel(data, "test.xlsx")

    workbook.close()
