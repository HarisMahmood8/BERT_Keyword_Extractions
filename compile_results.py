import openpyxl
import roberta_model

# name = input("Enter name of company: ")
# num_quarters = int(input("Enter number of quarters: "))
filenames = ["Adidas_Q3_2021",
             "Adidas_Y_2021",
             "Adidas_Q1_2022",
             "Adidas_Q2_2022",
             "Adidas_Q3_2022",
             "Adidas_Y_2022",
             "Adidas_Q1_2023",
             "Adidas_Q2_2023"]
workbooks = []
quarters = []
company = filenames[0].split("_")[0]

# Open result spreadsheets
# for Q in range(1, num_quarters+1):
    # workbook = openpyxl.load_workbook(f"sentiment_results_{name}_Q{Q}.xlsx")
    # workbooks.append(workbook)

for name in filenames:
    print("Running sentiment on", name)
    roberta_model.run_sentiment(name)

    quarter = name.replace(f"{company}_", "")
    quarters.append(quarter)
    workbook = openpyxl.load_workbook(f'sentiment_results_{name}.xlsx')
    workbooks.append(workbook)

res_dict_categories = {}
res_dict_keywords = {}

# Iterate through each sheet
for quarter in range(len(workbooks)):
    workbook = workbooks[quarter]
    sub_dict_categories = {}
    sub_dict_keywords = {}
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
        # List of categories found in the row's paragraph
        categories = row[3].split(", ")
        keywords = row[2].split(", ")

        # Get list of sentiment values assigned to each category
        for category in categories:
            key, value = category, float(row[1])

            if key not in sub_dict_categories.keys():
                sub_dict_categories[key] = []
            sub_dict_categories[key].append(value)

        # Now get sentiment values assigned to each keyword
        for keyword in keywords:
            key, value = keyword, float(row[1])

            if key not in sub_dict_keywords.keys():
                sub_dict_keywords[key] = []
            sub_dict_keywords[key].append(value)

    # Get average sentiment for each category/keyword and add them to the result dictionary
    for category, values in sub_dict_categories.items():
        average = sum(values) / len(values)

        if category not in res_dict_categories.keys():
            res_dict_categories[category] = []
            for Q in range(1, len(filenames)+1):
                res_dict_categories[category].append([])
        res_dict_categories[category][quarter].append(average)

    for keyword, values in sub_dict_keywords.items():
        average = sum(values) / len(values)

        if keyword not in res_dict_keywords:
            res_dict_keywords[keyword] = []
            for Q in range(1, len(filenames)+1):
                res_dict_keywords[keyword].append([])
        res_dict_keywords[keyword][quarter].append(average)

    workbook.close()

# Export results to spreadsheet
file_name = f"sentiment_results_{company}.xlsx"
res_workbook = openpyxl.Workbook()
sheet = res_workbook.active
sheet.title = f"{company} Results"

category_header = ["Category"]
for Q in quarters:
    category_header.append(Q)
sheet.append(category_header)
for category, scores in res_dict_categories.items():
    res_row = [category]
    for score in scores:
        if len(score):
            res_row.append(score[0])
        else:
            res_row.append("")
    sheet.append(res_row)

sheet.append([""])

keyword_header = ["Keyword"]
for Q in quarters:
    keyword_header.append(Q)
sheet.append(keyword_header)
for keyword, scores in res_dict_keywords.items():
    res_row = [keyword]
    for score in scores:
        if len(score):
            res_row.append(score[0])
        else:
            res_row.append("")
    sheet.append(res_row)

res_workbook.save(file_name)

print(f"Results compiled in {file_name}")
